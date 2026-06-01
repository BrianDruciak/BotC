"""Generate a synthetic games sheet to validate the importer end-to-end.

Uses scraped icon PNGs as stand-in "grimoire images" (the stub extractor ignores
pixels, so this exercises column detection, hyperlinks, winner parsing, and commit).
"""

import os

from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
IMG = os.path.join(ROOT, "data", "images")
OUT = os.path.join(ROOT, "data", "sample_games.xlsx")

wb = Workbook()
ws = wb.active
ws.title = "Games"
ws.append(["Game ID", "Date Played", "Grimoire Image", "Winner", "Script"])

rows = [
    (1, "2026-05-01", "Icon_imp.png", "Good", "Trouble Brewing"),
    (2, "2026-05-02", "Icon_fortune_teller.png", "evil", "Trouble Brewing"),
    (3, "2026-05-03", "Icon_po.png", "Good wins", "Bad Moon Rising"),
    (4, "2026-05-04", "Icon_vortox.png", "Evil team", "Sects & Violets"),
    (5, "2026-05-05", "Icon_baron.png", "good", ""),
]
for gid, date, icon, winner, script in rows:
    ws.append([gid, date, os.path.join(IMG, icon), winner, script])

# Row 6: image as an embedded hyperlink (display text differs from target).
ws.append([6, "2026-05-06", "see screenshot", "evil", "Trouble Brewing"])
cell = ws.cell(row=ws.max_row, column=3)
cell.hyperlink = Hyperlink(ref=cell.coordinate, target=os.path.join(IMG, "Icon_chef.png"))

wb.save(OUT)
print("wrote", OUT)
