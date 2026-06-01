"""
Excel batch importer.

Reads a games sheet where each row has link(s) to a final grimoire image and a
"good/evil won" note, then runs each row through the extraction + resolve + commit
pipeline. Columns are auto-detected (with an `inspect` mode to show what was found);
image links may be plain URLs, embedded Excel hyperlinks, or local file paths.
"""

from __future__ import annotations

import hashlib
import os
import re
import urllib.parse
import urllib.request
from dataclasses import dataclass, field

from openpyxl import load_workbook

from . import db, ingest
from .extract.base import Extractor, get_extractor

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
INBOX = os.path.join(ROOT, "data", "inbox")
USER_AGENT = "BotC-Stats-Agent/1.0"

IMAGE_HINTS = ("image", "img", "link", "url", "grim", "screenshot", "picture", "photo")
WINNER_HINTS = ("winner", "win", "result", "outcome", "won", "victor")
ID_HINTS = ("id", "game", "#", "no", "number")
DATE_HINTS = ("date", "played", "when", "day")
SCRIPT_HINTS = ("script", "edition", "module")

_GOOD = {"good", "g", "town", "townsfolk", "good win", "good wins", "good team", "village"}
_EVIL = {"evil", "e", "demon", "evil win", "evil wins", "evil team", "minions"}


@dataclass
class ColumnMap:
    image: list[int] = field(default_factory=list)
    winner: int | None = None
    external_id: int | None = None
    played_at: int | None = None
    script: int | None = None
    header_row: int = 1


def _norm_header(s) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(s or "").lower()).strip()


def _find_header_row(ws) -> int:
    for r in range(1, min(ws.max_row, 10) + 1):
        non_empty = sum(1 for c in ws[r] if c.value not in (None, ""))
        if non_empty >= 2:
            return r
    return 1


def detect_columns(ws) -> ColumnMap:
    cm = ColumnMap(header_row=_find_header_row(ws))
    headers = ws[cm.header_row]
    for idx, cell in enumerate(headers):
        h = _norm_header(cell.value)
        if not h:
            continue
        if any(k in h for k in WINNER_HINTS) and cm.winner is None:
            cm.winner = idx
        elif any(k in h for k in IMAGE_HINTS):
            cm.image.append(idx)
        elif any(k in h for k in DATE_HINTS) and cm.played_at is None:
            cm.played_at = idx
        elif any(k in h for k in SCRIPT_HINTS) and cm.script is None:
            cm.script = idx
        elif any(k in h for k in ID_HINTS) and cm.external_id is None:
            cm.external_id = idx
    return cm


def inspect(path: str) -> dict:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    cm = detect_columns(ws)
    headers = [(_norm_header(c.value), c.value) for c in ws[cm.header_row]]
    sample = []
    for r in range(cm.header_row + 1, min(cm.header_row + 4, ws.max_row + 1)):
        row = []
        for c in ws[r]:
            link = c.hyperlink.target if c.hyperlink else None
            row.append({"value": c.value, "hyperlink": link})
        sample.append(row)
    return {
        "sheet": ws.title,
        "dimensions": ws.dimensions,
        "header_row": cm.header_row,
        "headers": [{"index": i, "norm": h[0], "raw": h[1]} for i, h in enumerate(headers)],
        "detected": {
            "image_columns": cm.image,
            "winner_column": cm.winner,
            "id_column": cm.external_id,
            "date_column": cm.played_at,
            "script_column": cm.script,
        },
        "sample_rows": sample,
    }


def _normalize_winner(value) -> str | None:
    s = _norm_header(value)
    if not s:
        return None
    if s in _GOOD or s.startswith("good"):
        return "good"
    if s in _EVIL or s.startswith("evil"):
        return "evil"
    return None


def _looks_like_link(tok: str) -> bool:
    t = tok.strip().strip('"').strip("'")
    if not t:
        return False
    return (
        t.startswith("http")
        or os.path.sep in t
        or "/" in t
        or t.lower().endswith((".png", ".jpg", ".jpeg", ".webp"))
        or os.path.exists(t)
    )


def _cell_links(cell) -> list[str]:
    links: list[str] = []
    if cell.hyperlink and cell.hyperlink.target:
        links.append(cell.hyperlink.target)
    val = cell.value
    if isinstance(val, str):
        stripped = val.strip().strip('"').strip("'")
        # A single path may contain spaces, so don't split on spaces. Only treat
        # newlines / commas / semicolons as separators between multiple links.
        if os.path.exists(stripped):
            links.append(stripped)
        else:
            for tok in re.split(r"[\n,;]+", stripped):
                tok = tok.strip().strip('"').strip("'")
                if _looks_like_link(tok):
                    links.append(tok)
    seen, out = set(), []
    for l in links:
        if l not in seen:
            seen.add(l)
            out.append(l)
    return out


def _fetch_image(link: str) -> bytes | None:
    os.makedirs(INBOX, exist_ok=True)
    try:
        if link.startswith("http"):
            cache = os.path.join(INBOX, hashlib.sha256(link.encode()).hexdigest()[:16] + _ext(link))
            if os.path.exists(cache):
                with open(cache, "rb") as fh:
                    return fh.read()
            req = urllib.request.Request(link, headers={"User-Agent": USER_AGENT})
            with urllib.request.urlopen(req, timeout=60) as resp:
                blob = resp.read()
            with open(cache, "wb") as fh:
                fh.write(blob)
            return blob
        if os.path.exists(link):
            with open(link, "rb") as fh:
                return fh.read()
    except Exception as exc:  # noqa: BLE001
        print(f"    ! failed to fetch {link}: {exc}")
    return None


def _ext(link: str) -> str:
    path = urllib.parse.urlparse(link).path
    e = os.path.splitext(path)[1].lower()
    return e if e in (".png", ".jpg", ".jpeg", ".webp") else ".png"


def import_sheet(
    path: str,
    db_path: str = db.DEFAULT_DB,
    extractor: Extractor | None = None,
    commit_clean_only: bool = False,
    limit: int | None = None,
) -> dict:
    """Process every data row. Returns a summary dict."""
    db.init_db(db_path)
    extractor = extractor or get_extractor()
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    cm = detect_columns(ws)
    if not cm.image:
        raise ValueError("No image column detected. Run `inspect` and rename a column to include 'image'/'link'/'url'.")
    if cm.winner is None:
        raise ValueError("No winner column detected. Add a column with 'winner'/'result' containing good/evil.")

    summary = {"rows": 0, "committed": 0, "skipped_review": 0, "errors": 0, "no_image": 0, "game_ids": []}
    start = cm.header_row + 1
    for r in range(start, ws.max_row + 1):
        row = ws[r]
        if all(c.value in (None, "") for c in row):
            continue
        summary["rows"] += 1
        if limit and summary["rows"] > limit:
            summary["rows"] -= 1
            break

        winner = _normalize_winner(row[cm.winner].value) if cm.winner < len(row) else None
        links: list[str] = []
        for ci in cm.image:
            if ci < len(row):
                links.extend(_cell_links(row[ci]))
        if not links:
            summary["no_image"] += 1
            print(f"  row {r}: no image link, skipped")
            continue

        images = [b for b in (_fetch_image(l) for l in links) if b]
        if not images:
            summary["no_image"] += 1
            print(f"  row {r}: image(s) could not be fetched, skipped")
            continue

        try:
            extraction = extractor.extract(images, winner_hint=winner)
            resolved = ingest.resolve(extraction)
            resolved.source_images = links
            resolved.external_id = str(row[cm.external_id].value) if cm.external_id is not None and cm.external_id < len(row) and row[cm.external_id].value is not None else None
            resolved.played_at = str(row[cm.played_at].value) if cm.played_at is not None and cm.played_at < len(row) and row[cm.played_at].value is not None else None
            if cm.script is not None and cm.script < len(row) and row[cm.script].value:
                resolved.script = str(row[cm.script].value)

            if resolved.winner not in ("good", "evil"):
                summary["errors"] += 1
                print(f"  row {r}: could not determine winner, skipped")
                continue
            if commit_clean_only and resolved.needs_review:
                summary["skipped_review"] += 1
                print(f"  row {r}: needs review, not committed (use review UI)")
                continue

            gid = ingest.commit(resolved, db_path)
            summary["committed"] += 1
            summary["game_ids"].append(gid)
            flag = " [needs review]" if resolved.needs_review else ""
            print(f"  row {r}: committed game #{gid} ({resolved.player_count} seats, {resolved.winner} won){flag}")
        except Exception as exc:  # noqa: BLE001
            summary["errors"] += 1
            print(f"  row {r}: ERROR {exc}")

    return summary
