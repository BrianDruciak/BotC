"""
Command-line interface for the BOTC win-rate agent.

  python -m app.cli init-db
  python -m app.cli inspect data/games.xlsx
  python -m app.cli import  data/games.xlsx [--commit-clean-only] [--limit N] [--extractor gemini|stub]
  python -m app.cli demo    [--games 25]          # populate with stub data to test the pipeline
  python -m app.cli stats   [--min-games 1]
  python -m app.cli merge-players --keep NAME --merge NAME [NAME ...]
  python -m app.cli merge-players --apply-ocr
  python -m app.cli merge-players --suggest
  python -m app.cli reconcile                       # merge OCR aliases + align names to manual sheet
  python -m app.cli apply-manual                    # purge non-sheet players; player stats from xlsx
  python -m app.cli reset                          # delete the games DB
"""

from __future__ import annotations

import argparse
import json
import os

from . import db, ingest, stats
from .character_fix import fix_singleton_demon_duplicates
from .extract.base import get_extractor
from . import players as player_ops


def cmd_init_db(args):
    db.init_db(args.db)
    print(f"Initialized DB at {args.db}")


def cmd_inspect(args):
    from .importer import inspect
    print(json.dumps(inspect(args.path), indent=2, default=str))


def cmd_import(args):
    from .importer import import_sheet
    extractor = get_extractor(args.extractor) if args.extractor else None
    summary = import_sheet(
        args.path,
        db_path=args.db,
        extractor=extractor,
        commit_clean_only=args.commit_clean_only,
        limit=args.limit,
    )
    print("\nImport summary:", json.dumps(summary, indent=2))


def cmd_demo(args):
    db.init_db(args.db)
    extractor = get_extractor("stub")
    committed = 0
    for i in range(args.games):
        extraction = extractor.extract([f"demo-game-{i}".encode()])
        resolved = ingest.resolve(extraction)
        resolved.external_id = f"demo-{i+1}"
        ingest.commit(resolved, args.db)
        committed += 1
    print(f"Committed {committed} demo games to {args.db}")
    _print_stats(args.db, min_games=1)


def cmd_stats(args):
    _print_stats(args.db, min_games=args.min_games)


def cmd_merge_players(args):
    if args.apply_ocr:
        results = player_ops.apply_ocr_merges(args.db)
        if not results:
            print("No OCR merges applied (nothing to merge).")
            return
        print(f"Applied {len(results)} merge group(s):\n")
        for r in results:
            merged = ", ".join(f"{m['name']!r} ({m['seats']} seats)" for m in r["merged"])
            print(f"  kept {r['keep_name']!r} <- {merged}")
        _print_stats(args.db, min_games=1)
        return

    if args.suggest:
        pairs = player_ops.suggest_merges(args.db)
        if not pairs:
            print("No likely duplicates found.")
            return
        print(f"Likely OCR duplicates ({len(pairs)} pairs):\n")
        for p in pairs[:30]:
            print(
                f"  {p['ratio']:5.1f}  {p['a']!r} ({p['a_seats']})"
                f"  <->  {p['b']!r} ({p['b_seats']})"
            )
        return

    if not args.keep or not args.merge:
        raise SystemExit("Use --keep and --merge, or --apply-ocr, or --suggest.")
    result = player_ops.merge_by_names(args.db, args.keep, args.merge)
    merged = ", ".join(f"{m['name']!r}" for m in result["merged"]) or "(none)"
    print(f"Merged into {result['keep_name']!r}: {merged} ({result['seats_moved']} seats moved)")


def cmd_reconcile(args):
    result = player_ops.reconcile_manual(args.db)
    merges = result["merges"]
    if merges:
        print(f"Applied {len(merges)} merge group(s):")
        for r in merges:
            merged = ", ".join(f"{m['name']!r} ({m['seats']})" for m in r["merged"])
            print(f"  kept {r['keep_name']!r} <- {merged}")
    else:
        print("No new merges needed.")
    if result["renames"]:
        print("\nRenamed:")
        for line in result["renames"]:
            print(f"  {line}")
    _compare_manual(args.db)
    _print_stats(args.db, min_games=1)


def cmd_reset(args):
    if os.path.exists(args.db):
        os.remove(args.db)
        print(f"Deleted {args.db}")
    else:
        print("No DB to delete.")


def cmd_apply_manual(args):
    purge = player_ops.purge_non_manual_players(args.db)
    print(
        f"Removed {purge['removed_players']} OCR-only players "
        f"({purge['removed_seats']} hallucinated seats)."
    )
    demon_fixes = fix_singleton_demon_duplicates(args.db)
    if demon_fixes:
        print("\nFixed duplicate demon tags:")
        for f in demon_fixes:
            print(
                f"  game #{f['game_id']} {f['demon']}: kept {f['kept']!r}, "
                f"demoted {f['demoted']}"
            )
    print(f"\nPlayer stats now served from stats (1).xlsx ({purge.get('allowed', 43)} players).")
    _print_stats(args.db, min_games=1)


def _compare_manual(db_path):
    from openpyxl import load_workbook

    path = os.path.join(os.path.dirname(db_path), "..", "stats (1).xlsx")
    path = os.path.normpath(path)
    if not os.path.exists(path):
        path = os.path.join(os.path.dirname(db_path), "stats (1).xlsx")
    if not os.path.exists(path):
        return

    wb = load_workbook(path, data_only=True)
    ws = wb["Personal Winrates"]
    manual = {}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        manual[str(name).strip()] = {
            "games": int(ws.cell(r, 4).value or 0),
            "wins": int(ws.cell(r, 2).value or 0),
        }

    db_players = {p["name"]: p for p in stats.player_stats(db_path, min_games=1)}
    mismatches = []
    for name, m in manual.items():
        p = db_players.get(name)
        if not p:
            mismatches.append(f"  {name}: missing in DB")
        elif p["games"] != m["games"] or p["wins"] != m["wins"]:
            mismatches.append(
                f"  {name}: manual {m['games']}G/{m['wins']}W"
                f" vs DB {p['games']}G/{p['wins']}W"
            )

    print(f"\n=== vs stats (1).xlsx ({len(manual)} players) ===")
    if mismatches:
        print(f"Still off ({len(mismatches)}):")
        for line in mismatches[:20]:
            print(line)
    else:
        print("All player game/win counts match manual sheet.")


def _print_stats(db_path, min_games=0):
    ov = stats.overview(db_path)
    print("\n=== Overview ===")
    print(f"  games={ov['games']}  seats={ov['seats']}  players={ov['players']}"
          f"  source={ov.get('player_source', 'database')}")
    print(f"  good wins={ov['good_wins']} ({ov['good_win_pct']}%)  evil wins={ov['evil_wins']}")
    print(f"  games needing review: {ov['games_need_review']}")

    chars = stats.character_stats(db_path, min_games=min_games)
    print(f"\n=== Top characters by win% (min {min_games} games) ===")
    print(f"  {'character':22} {'type':10} {'G':>3} {'W':>3} {'win%':>6}")
    for c in chars[:15]:
        print(f"  {c['name']:22} {c['type']:10} {c['games']:>3} {c['wins']:>3} {c['win_pct']:>6}")

    players = stats.player_stats(db_path, min_games=min_games)
    print(f"\n=== Players by win% (min {min_games} games) ===")
    print(f"  {'player':16} {'G':>3} {'W':>3} {'win%':>6}  good%  evil%")
    for p in players[:15]:
        gp = "-" if p["good_win_pct"] is None else p["good_win_pct"]
        ep = "-" if p["evil_win_pct"] is None else p["evil_win_pct"]
        print(f"  {p['name']:16} {p['games']:>3} {p['wins']:>3} {p['win_pct']:>6}  {gp:>5}  {ep:>5}")


def build_parser():
    p = argparse.ArgumentParser(prog="botc", description="BOTC win-rate agent")
    p.add_argument("--db", default=db.DEFAULT_DB, help="SQLite DB path")
    sub = p.add_subparsers(dest="cmd", required=True)

    sub.add_parser("init-db").set_defaults(func=cmd_init_db)

    sp = sub.add_parser("inspect")
    sp.add_argument("path")
    sp.set_defaults(func=cmd_inspect)

    sp = sub.add_parser("import")
    sp.add_argument("path")
    sp.add_argument("--extractor", choices=["gemini", "stub"], default=None)
    sp.add_argument("--commit-clean-only", action="store_true")
    sp.add_argument("--limit", type=int, default=None)
    sp.set_defaults(func=cmd_import)

    sp = sub.add_parser("demo")
    sp.add_argument("--games", type=int, default=25)
    sp.set_defaults(func=cmd_demo)

    sp = sub.add_parser("stats")
    sp.add_argument("--min-games", type=int, default=0)
    sp.set_defaults(func=cmd_stats)

    sp = sub.add_parser("merge-players")
    sp.add_argument("--keep", help="Canonical player name to keep")
    sp.add_argument("--merge", nargs="+", help="Duplicate name(s) to fold into --keep")
    sp.add_argument("--apply-ocr", action="store_true", help="Apply built-in OCR merge groups")
    sp.add_argument("--suggest", action="store_true", help="List likely duplicate pairs")
    sp.set_defaults(func=cmd_merge_players)

    sub.add_parser("reconcile").set_defaults(func=cmd_reconcile)

    sub.add_parser("apply-manual").set_defaults(func=cmd_apply_manual)

    sub.add_parser("reset").set_defaults(func=cmd_reset)
    return p


def main(argv=None):
    args = build_parser().parse_args(argv)
    args.func(args)


if __name__ == "__main__":
    main()
