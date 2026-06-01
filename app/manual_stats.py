"""
Authoritative player win rates from the manual Excel sheet.

When stats (1).xlsx is present, player_stats() serves this file instead of
aggregating OCR game_seat rows (which include hallucinated names/seats).
"""

from __future__ import annotations

import os

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DEFAULT_XLSX = os.path.join(ROOT, "stats (1).xlsx")


def manual_xlsx_path() -> str:
    return os.environ.get("MANUAL_STATS_XLSX", DEFAULT_XLSX)


def load_manual_players(min_games: int = 0) -> list[dict] | None:
    path = manual_xlsx_path()
    if not os.path.exists(path):
        return None

    wb = load_workbook(path, data_only=True)
    ws = wb["Personal Winrates"] if "Personal Winrates" in wb.sheetnames else wb.active

    out: list[dict] = []
    for row_idx, r in enumerate(range(2, ws.max_row + 1), start=1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        name = str(name).strip()
        wins = int(ws.cell(r, 2).value or 0)
        losses = int(ws.cell(r, 3).value or 0)
        games = int(ws.cell(r, 4).value or 0)
        good_wins = int(ws.cell(r, 6).value or 0)
        good_games = int(ws.cell(r, 8).value or 0)
        evil_wins = int(ws.cell(r, 10).value or 0)
        evil_games = int(ws.cell(r, 12).value or 0)

        if games < min_games:
            continue

        out.append(
            {
                "id": row_idx,
                "name": name,
                "games": games,
                "wins": wins,
                "losses": losses,
                "win_pct": round(100.0 * wins / games, 1) if games else 0.0,
                "good_games": good_games,
                "good_win_pct": round(100.0 * good_wins / good_games, 1) if good_games else None,
                "evil_games": evil_games,
                "evil_win_pct": round(100.0 * evil_wins / evil_games, 1) if evil_games else None,
                "source": "manual",
            }
        )

    out.sort(key=lambda d: (-d["win_pct"], -d["games"], d["name"]))
    return out


def manual_player_names() -> set[str] | None:
    players = load_manual_players(min_games=0)
    if players is None:
        return None
    return {p["name"] for p in players}


def get_manual_player(player_id: int) -> dict | None:
    players = load_manual_players(min_games=0)
    if not players:
        return None
    for p in players:
        if p["id"] == player_id:
            return p
    return None
